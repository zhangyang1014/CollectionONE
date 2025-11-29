#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成功能实现详细报告
"""

import openpyxl
from pathlib import Path
from collections import defaultdict

# 获取Excel文件路径
script_dir = Path(__file__).parent
project_root = script_dir.parent
excel_path = project_root / "PRD需求文档" / "CCO 系统功能设计.xlsx"

# 加载Excel
wb = openpyxl.load_workbook(str(excel_path))
ws = wb["控台功能点"]

# 按模块分类统计
modules = defaultdict(lambda: {
    "已实现": [],
    "部分实现": [],
    "未实现": [],
    "后端Mock": [],
    "待定": []
})

current_module = None
for row_idx in range(2, ws.max_row + 1):
    module = ws.cell(row_idx, 1).value
    category1 = ws.cell(row_idx, 2).value
    category2 = ws.cell(row_idx, 3).value
    desc = ws.cell(row_idx, 4).value
    diff = ws.cell(row_idx, 7).value
    
    if not desc or str(desc).strip() == "":
        continue
    
    # 确定模块
    if module and str(module).strip():
        current_module = str(module).strip()
    elif category1 and str(category1).strip() and "管理" in str(category1):
        current_module = str(category1).strip()
    
    if not current_module:
        current_module = "其他"
    
    # 提取状态
    status = "未知"
    if diff:
        diff_str = str(diff)
        if "已实现" in diff_str:
            status = "已实现"
        elif "部分实现" in diff_str:
            status = "部分实现"
        elif "未实现" in diff_str:
            status = "未实现"
        elif "后端Mock" in diff_str:
            status = "后端Mock"
        elif "待定" in diff_str:
            status = "待定"
    
    # 添加到对应模块
    feature_info = {
        'desc': desc,
        'category': f"{category1 or ''} > {category2 or ''}".strip(' >'),
        'note': diff
    }
    
    if status in modules[current_module]:
        modules[current_module][status].append(feature_info)

# 生成报告
print("\n" + "="*100)
print("CCO 控台功能实现情况详细报告")
print("="*100)

for module_name in sorted(modules.keys()):
    module_data = modules[module_name]
    total = sum(len(module_data[status]) for status in module_data)
    
    if total == 0:
        continue
    
    implemented = len(module_data["已实现"]) + len(module_data["部分实现"]) + len(module_data["后端Mock"])
    completion_rate = implemented / total * 100
    
    print(f"\n【{module_name}】 - 共 {total} 个功能，完成率 {completion_rate:.1f}%")
    print("-" * 100)
    
    # 已实现
    if module_data["已实现"]:
        print(f"\n  ✅ 已实现 ({len(module_data['已实现'])}):")
        for feature in module_data["已实现"]:
            print(f"     • {feature['desc'][:60]}")
    
    # 部分实现
    if module_data["部分实现"]:
        print(f"\n  🟡 部分实现 ({len(module_data['部分实现'])}):")
        for feature in module_data["部分实现"]:
            print(f"     • {feature['desc'][:60]}")
    
    # 后端Mock
    if module_data["后端Mock"]:
        print(f"\n  🔵 后端Mock ({len(module_data['后端Mock'])}):")
        for feature in module_data["后端Mock"]:
            print(f"     • {feature['desc'][:60]}")
    
    # 未实现
    if module_data["未实现"]:
        print(f"\n  ❌ 未实现 ({len(module_data['未实现'])}):")
        for feature in module_data["未实现"]:
            print(f"     • {feature['desc'][:60]}")
    
    # 待定
    if module_data["待定"]:
        print(f"\n  ⚪ 待定 ({len(module_data['待定'])}):")
        for feature in module_data["待定"]:
            print(f"     • {feature['desc'][:60]}")

print("\n" + "="*100)
print("\n📝 详细说明已保存在Excel文件的G列")
print(f"📁 文件路径: {excel_path}\n")


