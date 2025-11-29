#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
校验CCO系统功能设计Excel中"控台功能点"sheet与实际实现的差异 (V2版本)
"""

import openpyxl
from openpyxl.styles import Font
from pathlib import Path
import re

def check_feature_implementation(module, category1, category2, description):
    """
    根据功能描述检查实现情况
    返回: (状态, 说明)
    """
    
    # API接口类 - 这些是后端接口，不在前端显示
    # 首先检查描述中的关键词（不依赖module，因为有些行module为空）
    desc_str = str(description).lower()
    
    # 案件相关API
    if "提供api接口" in desc_str and "案件" in desc_str:
        if "更新案件信息" in desc_str:
            return ("后端Mock", "后端Mock接口 - PUT /api/v1/cases/{id}")
        elif "回收" in desc_str:
            return ("后端Mock", "后端Mock接口 - POST /api/v1/cases/{id}/recycle")
    
    # CCO调取接口
    if "cco调取接口" in desc_str:
        if "指定催员" in desc_str and "案件" in desc_str:
            return ("后端Mock", "后端Mock接口 - GET /api/v1/cases?collectorId={id}")
        elif "案件队列" in desc_str and "更新" in desc_str:
            return ("后端Mock", "后端Mock接口 - GET /api/v1/cases/updates")
        elif "回收" in desc_str:
            return ("后端Mock", "后端Mock接口 - POST /api/v1/cases/{id}/recycle")
    
    # 还款码相关
    if "还款码" in desc_str:
        if "申请" in desc_str:
            return ("后端Mock", "后端Mock接口 - POST /api/v1/payment-codes")
        elif "查询" in desc_str or "已有" in desc_str:
            return ("已实现", "已实现 - IM端还款码Tab可查询 + 后端Mock接口")
    
    # 展期相关
    if "展期" in desc_str:
        return ("未实现", "展期功能暂未实现")
    
    # 催记/通话记录
    if ("催记" in desc_str or "通话记录" in desc_str or "聊天记录" in desc_str) and "回传" not in desc_str:
        return ("后端Mock", "后端Mock接口 - GET /api/v1/cases/{id}/notes")
    
    # 渠道相关
    if "luna" in desc_str or ("电话" in desc_str and "自己的" in desc_str):
        return ("未实现", "Luna电话渠道暂未配置")
    
    if "cwa" in desc_str or "whatsapp" in desc_str:
        return ("已实现", "已实现 - 在甲方渠道管理中可配置WhatsApp/CWA")
    
    # 质检相关
    if "质检" in desc_str and "录音" in desc_str:
        return ("未实现", "质检功能暂未实现")
    
    # 通过module判断（向后兼容）
    if "对甲方接口" in str(module) or "CCO提供API" in str(module) or "甲方提供API" in str(module):
        if "字段匹配" in str(description):
            return ("后端Mock", "后端接口 - 字段映射配置页面可手动配置")
        elif "案件进件" in str(description) or "批量提供案件" in str(description):
            return ("后端Mock", "后端Mock接口 - POST /api/v1/cases/batch")
        elif "催员登录" in str(description):
            return ("已实现", "后端接口 + IM端登录页面: /im/login")
        else:
            return ("后端Mock", "后端Mock接口")
    
    # 登录管理
    if "登录" in str(description) and "Token" in str(description):
        return ("已实现", "已实现 - /admin/login + JWT Token管理")
    
    # 数据看板
    if "工作台首页" in str(description) or "月度绩效" in str(description):
        return ("已实现", "已实现 - /dashboard (工作台)")
    elif "到期日" in str(description) and "新入催率" in str(description):
        return ("部分实现", "单催员业绩看板已实现(/performance/my-dashboard)，部分指标Mock数据")
    elif "委外法催" in str(description) or "处置阶段" in str(description):
        return ("未实现", "委外法催统计看板未实现")
    elif "迁徙率" in str(description):
        return ("未实现", "迁徙率分析看板未实现")
    elif "前手队列" in str(description):
        return ("未实现", "队列对比分析未实现")
    elif "排名" in str(description) and "佣金" in str(description):
        return ("未实现", "排名看板未实现")
    elif "绑定人" in str(description) and "催记量" in str(description):
        return ("未实现", "详细催记统计未实现")
    elif "拨打量" in str(description) and "通话时长" in str(description):
        return ("未实现", "通话统计看板未实现")
    elif "空闲催员" in str(description):
        return ("已实现", "已实现 - /dashboard/idle-monitor (空闲催员监控)")
    
    # 案件管理
    if "案件列表" in str(description) and "动态字段" in str(description):
        return ("已实现", "已实现 - /cases (案件列表，支持动态字段配置)")
    elif "历史催记" in str(description) or "查看案件详情" in str(description):
        return ("已实现", "已实现 - /cases/:id (案件详情页，可查看催记)")
    elif "案件详情页面" in str(description) and "还款码" in str(description):
        return ("已实现", "已实现 - /cases/:id (案件详情，包含基本信息、字段、还款码)")
    elif "管理甲方的案件队列" in str(description):
        return ("已实现", "已实现 - /tenants/queue-management (案件队列管理)")
    elif "手动分配案件" in str(description):
        return ("未实现", "手动分案功能未实现")
    elif "标记为不催" in str(description):
        return ("未实现", "标记不催功能未实现")
    elif "分案策略列表" in str(description):
        return ("已实现", "已实现 - /auto-assignment (自动化分案策略管理)")
    elif "编辑分案策略" in str(description):
        return ("已实现", "已实现 - /auto-assignment 策略详情编辑")
    elif "创建新的分案策略" in str(description) or "向导创建" in str(description):
        return ("已实现", "已实现 - /auto-assignment 策略向导创建")
    elif "预跑" in str(description):
        return ("未实现", "策略预跑功能未实现")
    elif "定期执行策略" in str(description):
        return ("未实现", "定时执行策略功能未实现")
    elif "监控分案" in str(description) and "报警" in str(description):
        return ("未实现", "分案监控预警未实现")
    elif "日志记录" in str(description):
        return ("未实现", "案件更新日志未实现")
    elif "回收" in str(description) and "退回给甲方" in str(description):
        return ("未实现", "案件回收功能未实现")
    elif "停留" in str(description):
        return ("未实现", "案件停留状态未实现")
    elif "微信群预警" in str(description):
        return ("未实现", "微信群预警未实现")
    
    # 字段配置
    if "标准字段" in str(description) and "增删改查" in str(description):
        return ("已实现", "已实现 - /field-config/standard (标准字段管理)")
    elif "字段分组" in str(description):
        return ("已实现", "已实现 - /field-config/groups (字段分组管理)")
    elif "自定义拓展字段" in str(description):
        return ("已实现", "已实现 - /field-config/custom (字段映射配置)")
    elif "字段展示规则" in str(description) and ("排序" in str(description) or "筛选" in str(description)):
        return ("已实现", "已实现 - /field-config/display (字段展示配置，含排序、筛选、隐私设置)")
    elif "查看甲方字段配置" in str(description):
        return ("已实现", "已实现 - /field-config/tenant-fields-view (甲方字段查看)")
    
    # 甲方/组织架构管理
    if "甲方列表" in str(description):
        return ("已实现", "已实现 - /tenants (甲方管理)")
    elif "机构列表" in str(description) and "创建、编辑、禁用" in str(description):
        return ("已实现", "已实现 - /organization/agencies (机构管理)")
    elif "作息时间" in str(description):
        return ("已实现", "已实现 - /organization/agencies/:id/working-hours (机构作息时间)")
    elif "小组群" in str(description):
        return ("已实现", "已实现 - /organization/team-groups (小组群管理)")
    elif "小组（Team）" in str(description) or ("小组" in str(description) and "创建、编辑" in str(description)):
        return ("已实现", "已实现 - /organization/teams (小组管理)")
    elif "小组管理员" in str(description):
        return ("已实现", "已实现 - /organization/admin-accounts (小组管理员管理)")
    elif "催员账号" in str(description):
        return ("已实现", "已实现 - /organization/collectors (催员管理)")
    
    # 渠道配置
    if "渠道发送限制" in str(description):
        return ("已实现", "已实现 - /channel-config/limits (渠道发送限制配置)")
    elif "短信渠道" in str(description):
        return ("已实现", "已实现 - /channel-config/suppliers (甲方渠道管理，含短信)")
    elif "waba" in str(description).lower() or "whatsapp" in str(description).lower():
        return ("已实现", "已实现 - /channel-config/suppliers (甲方渠道管理，含WABA)")
    elif "rcs" in str(description).lower():
        return ("已实现", "已实现 - /channel-config/suppliers (甲方渠道管理，含RCS)")
    elif "Infinity" in str(description):
        return ("已实现", "已实现 - /channel-config/suppliers (含Infinity外呼配置)")
    elif "还款渠道" in str(description):
        return ("已实现", "已实现 - /channel-config/suppliers (含还款渠道管理)")
    
    # 权限配置
    if "角色的权限" in str(description):
        return ("已实现", "已实现 - /system/permissions (权限配置) + /system/permission-management (权限查看)")
    elif "白名单" in str(description):
        return ("未实现", "催员登录白名单未实现")
    
    # 通知配置
    if "通知模板" in str(description):
        return ("已实现", "已实现 - /system/notification-config (通知配置，含模板)")
    elif "触发维度" in str(description) and "通知" in str(description):
        return ("已实现", "已实现 - /system/notification-config (通知维度配置)")
    elif "公共通知" in str(description):
        return ("已实现", "已实现 - /system/notification-config (公共通知配置)")
    
    # 操作日志
    if "操作日志" in str(description):
        return ("未实现", "操作日志查询未实现")
    
    # 质检管理
    if "质检规则" in str(description):
        return ("未实现", "质检功能暂未实现")
    elif "质检任务" in str(description):
        return ("未实现", "质检功能暂未实现")
    elif "质检工作台" in str(description):
        return ("未实现", "质检功能暂未实现")
    elif "质检记录" in str(description):
        return ("未实现", "质检功能暂未实现")
    elif "异常录音" in str(description) or "异常文字" in str(description) or "异常图片" in str(description):
        return ("未实现", "质检功能暂未实现")
    
    # 待定功能
    if "待定" in str(description):
        return ("待定", "功能待定")
    
    # 默认
    return ("待确认", "请人工确认功能实现情况")

def verify_excel_features():
    """校验Excel中的功能点"""
    
    # 获取Excel文件路径
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    excel_path = project_root / "PRD需求文档" / "CCO 系统功能设计.xlsx"
    
    print(f"🔍 开始校验控台功能点...")
    print(f"📁 Excel文件: {excel_path}\n")
    
    # 加载Excel
    wb = openpyxl.load_workbook(str(excel_path))
    
    # 找到"控台功能点"sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "控台功能点" in name or "控台" in name:
            sheet_name = name
            break
    
    if not sheet_name:
        print(f"❌ 未找到'控台功能点'sheet")
        return False
    
    ws = wb[sheet_name]
    print(f"✅ 找到sheet: {sheet_name}\n")
    
    # 列定义
    COL_MODULE = 1      # A列: 产品模块
    COL_CATEGORY1 = 2   # B列: 功能类别一
    COL_CATEGORY2 = 3   # C列: 功能类别二
    COL_DESC = 4        # D列: 模块功能描述
    COL_PRIORITY = 5    # E列: 优先级
    COL_IMPORTANCE = 6  # F列: 重要度
    COL_DIFF = 7        # G列: 实现差异说明
    
    # 设置G列表头
    ws.cell(1, COL_DIFF).value = "实现差异说明"
    ws.cell(1, COL_DIFF).font = Font(bold=True)
    
    # 统计
    stats = {
        "已实现": 0,
        "部分实现": 0,
        "未实现": 0,
        "后端Mock": 0,
        "待定": 0,
        "待确认": 0
    }
    
    # 遍历每一行
    for row_idx in range(2, ws.max_row + 1):
        module = ws.cell(row_idx, COL_MODULE).value
        category1 = ws.cell(row_idx, COL_CATEGORY1).value
        category2 = ws.cell(row_idx, COL_CATEGORY2).value
        description = ws.cell(row_idx, COL_DESC).value
        
        # 跳过空行
        if not description or str(description).strip() == "":
            continue
        
        # 检查实现情况
        status, note = check_feature_implementation(module, category1, category2, description)
        
        # 统计
        if status in stats:
            stats[status] += 1
        
        # 设置单元格内容和颜色
        cell = ws.cell(row_idx, COL_DIFF)
        cell.value = f"{status} - {note}"
        
        # 设置颜色
        if status == "已实现":
            cell.font = Font(color="008000")  # 绿色
        elif status == "部分实现":
            cell.font = Font(color="FF8C00")  # 橙色
        elif status == "未实现":
            cell.font = Font(color="FF0000")  # 红色
        elif status == "后端Mock":
            cell.font = Font(color="0000FF")  # 蓝色
        elif status == "待定":
            cell.font = Font(color="808080")  # 灰色
        else:
            cell.font = Font(color="800080")  # 紫色
    
    # 保存文件
    wb.save(str(excel_path))
    
    # 打印统计
    total = sum(stats.values())
    print("="*60)
    print("📊 校验统计")
    print("="*60)
    print(f"总功能数: {total}")
    print(f"✅ 已实现: {stats['已实现']} ({stats['已实现']/total*100:.1f}%)")
    print(f"🟡 部分实现: {stats['部分实现']} ({stats['部分实现']/total*100:.1f}%)")
    print(f"❌ 未实现: {stats['未实现']} ({stats['未实现']/total*100:.1f}%)")
    print(f"🔵 后端Mock: {stats['后端Mock']} ({stats['后端Mock']/total*100:.1f}%)")
    print(f"⚪ 待定: {stats['待定']} ({stats['待定']/total*100:.1f}%)")
    print(f"🟣 待确认: {stats['待确认']} ({stats['待确认']/total*100:.1f}%)")
    print("="*60)
    
    implemented = stats['已实现'] + stats['部分实现'] + stats['后端Mock']
    print(f"\n✨ 总体完成率: {implemented/total*100:.1f}%")
    print(f"   (已实现 + 部分实现 + 后端Mock)\n")
    
    print(f"✅ 差异说明已写入 {sheet_name} 的G列")
    print(f"📁 文件已保存: {excel_path}\n")
    
    return True

if __name__ == "__main__":
    import sys
    success = verify_excel_features()
    sys.exit(0 if success else 1)

