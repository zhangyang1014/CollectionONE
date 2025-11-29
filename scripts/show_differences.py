#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
显示所有有差异的功能点
"""

import openpyxl
from pathlib import Path

# 获取Excel文件路径
script_dir = Path(__file__).parent
project_root = script_dir.parent
excel_path = project_root / "PRD需求文档" / "CCO 系统功能设计.xlsx"

# 加载Excel
wb = openpyxl.load_workbook(str(excel_path))
ws = wb["控台功能点"]

print("="*100)
print("🔍 实现差异详情")
print("="*100)

# 分类存储
no_diff = []
partial_diff = []
impl_diff = []
name_diff = []

for row_idx in range(2, ws.max_row + 1):
    desc = ws.cell(row_idx, 4).value  # D列
    status = ws.cell(row_idx, 7).value  # G列
    detail = ws.cell(row_idx, 8).value  # H列
    
    if not detail or not desc:
        continue
    
    detail_str = str(detail)
    
    item = {
        'row': row_idx,
        'desc': str(desc)[:80],
        'status': status,
        'detail': detail_str
    }
    
    if "⚠️ 部分差异" in detail_str:
        partial_diff.append(item)
    elif "⚠️ 实现方式不同" in detail_str:
        impl_diff.append(item)
    elif "⚠️ 命名差异" in detail_str:
        name_diff.append(item)

# 显示有差异的功能
print("\n" + "="*100)
print("⚠️ 实现方式不同的功能 ({} 个)".format(len(impl_diff)))
print("="*100)
for item in impl_diff:
    print(f"\n行{item['row']}: 【{item['status']}】")
    print(f"  需求描述: {item['desc']}")
    print(f"  差异说明: {item['detail']}")

print("\n" + "="*100)
print("⚠️ 部分差异的功能 ({} 个)".format(len(partial_diff)))
print("="*100)
for item in partial_diff:
    print(f"\n行{item['row']}: 【{item['status']}】")
    print(f"  需求描述: {item['desc']}")
    print(f"  差异说明: {item['detail']}")

print("\n" + "="*100)
print("⚠️ 命名差异的功能 ({} 个)".format(len(name_diff)))
print("="*100)
for item in name_diff:
    print(f"\n行{item['row']}: 【{item['status']}】")
    print(f"  需求描述: {item['desc']}")
    print(f"  差异说明: {item['detail']}")

print("\n" + "="*100)
print("📊 差异汇总")
print("="*100)
total_diff = len(impl_diff) + len(partial_diff) + len(name_diff)
print(f"实现方式不同: {len(impl_diff)}")
print(f"部分差异: {len(partial_diff)}")
print(f"命名差异: {len(name_diff)}")
print(f"总计: {total_diff} 个功能存在差异")
print("="*100)




