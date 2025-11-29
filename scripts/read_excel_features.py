#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取Excel中的功能点列表
"""

import openpyxl
from pathlib import Path

# 获取Excel文件路径
script_dir = Path(__file__).parent
project_root = script_dir.parent
excel_path = project_root / "PRD需求文档" / "CCO 系统功能设计.xlsx"

# 加载Excel
wb = openpyxl.load_workbook(str(excel_path))

# 找到"控台功能点"sheet
sheet_name = None
for name in wb.sheetnames:
    if "控台功能点" in name or "控台" in name:
        sheet_name = name
        break

if sheet_name:
    ws = wb[sheet_name]
    print(f"找到sheet: {sheet_name}")
    print(f"总行数: {ws.max_row}")
    print(f"总列数: {ws.max_column}")
    print("\n前10行数据:")
    print("="*100)
    
    # 打印前10行
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(8, ws.max_column + 1)):  # 只显示前7列
            cell_value = ws.cell(row_idx, col_idx).value
            row_data.append(str(cell_value or "")[:30])  # 限制每个单元格最多显示30字符
        print(f"行{row_idx}: {' | '.join(row_data)}")
    
    print("\n" + "="*100)
    print("\n所有功能点列表 (D列):")
    print("="*100)
    
    # 找到D列的所有功能点
    feature_col = 4  # D列是第4列
    features = []
    for row_idx in range(2, ws.max_row + 1):  # 从第2行开始(跳过表头)
        cell_value = ws.cell(row_idx, feature_col).value
        if cell_value and str(cell_value).strip() and str(cell_value).strip() != "None":
            features.append(str(cell_value).strip())
    
    # 按模块分组显示
    current_module = ""
    for idx, feature in enumerate(features, 1):
        # 如果是大类标题(没有具体操作描述)
        if len(feature) < 15 and ("管理" in feature or "配置" in feature or "看板" in feature):
            current_module = feature
            print(f"\n【{current_module}】")
        else:
            print(f"  {idx}. {feature}")
    
    print("\n" + "="*100)
    print(f"总计: {len(features)} 个功能点")
else:
    print("未找到'控台功能点'sheet")
    print(f"可用的sheet: {wb.sheetnames}")


